import datetime
import tkinter as tk
from tkinter import filedialog
import openpyxl
from tkinter import messagebox
from tkinter import ttk

gv_title = "SUREINFO Data Toolkit" 
gv_title_label = "SUREINFO SAP S/4 Data Migration Cleansing Tool"
gv_button_label = "Select Excel File"
gv_version = "Version 1.0.7"
gv_licesne_expire_date = datetime.datetime(2025, 12, 31)

# environment: python 3.9.7, openpyxl 3.0.9, tkinter 8.6
# to pack the application, use the following command:
# pyinstaller --onefile --noconsole --icon=favicon.ico --name 'SUREINFO SAP S4 Data Toolkit v1.0.7' .\SUREINFO_SAP_DATA_CHECK

# 2024-04-08 backlog v0.0.2:
# 1. create initial screen - done
# 2. a button on top left to open dialog and select the excel file.  - done
# 3. a file path label to the right side of the button, smaller size.  - done
# 4. after a file was selected, populate the unhiden sheet names as buttons with action data check -done

# 2024-04-09 backlogs v0.0.3:
# 5. show progress bar - done

# 2024-04-09 backlogs v0.0.4:
# 6. add a button to copy the whole result into clipboard after data checked - nok - not easy to achive. 
# 7. performance tunning, load one sheet each time, not all sheets - not possible, need to load all sheets to get the sheet names

# 2024-04-10 backlogs v0.0.5:
# 8. check the reference, if the reference is not found in the area, show the error message - done 
# 9. change the icon for app - done

# 2024-04-10 backlogs v0.0.6:
# 10. format the output row number in fomrat 1: [Date Month SAP]: [Error Messages] 
# 11. performance tunning using one single load workbook. - done (be careful with the readonly mode, it is not working)
# 12. add the license check, if the date is bigger than 2024-01-01, then show the lincese expired message, confirm and close the application - done 

# 2024-04-10 backlogs v0.0.7:
# 13. try to optimize, however code messed up, need to clean up. - done

# 2024-04-11 backlogs v0.0.8:

# 2024-04-11 backlogs v0.0.8:
# 14. bugfix of the cell value length check 
# 15. bugfix of reopen status update issue
# 16. bugfix of runtime errors 
# 17. fix the issue of filtered_check_list

# 2024-04-11 backlogs v0.0.9:
# 18. basic version can be used for work
# 19. change min_row to 3 to adapt different format of ref data

# 2024-04-11 backlogs v1.0.0:
# 20. publish for work use
# 21. add duplication check

# 2024-04-11 backlogs v1.0.1:
# 22. fixed the error for open with permission issue
# 23. add the button to copy the result to clipboard

# 2024-04-12 backlogs v1.0.2:
# 24. add more space to the button, and changed the background of the text 
# 25. put the copy result message from popup to the status text

# 2024-04-14 backlogs v1.0.4:
# 26. bug fix for the mandaotry check to consider None and space as well 

# 2024-04-15 backlogs v1.0.5:
# 27. bug fix for the breakpoint 
# 28. mandatory check for F (Fixed Value) as well 

# 2024-05-10 backlogs v1.0.6:
# 29. bug fix for null mandatory indicator

# 2024-05-10 backlogs v1.0.7:
# 30. bug fix for reference value check with data field and reset the check 

# 2025-05-05 going forward, the change log will be maintained in readme.md only

# TODO sum check 

# ------------------ code starts here ------------------

# global gv_file_path, sheet_frame, message_frame, status_frame, message_text, status_text, workbook 
# sheet_frame = None
# message_frame = None
# status_frame = None
# message_text = None
status_text = None # dont comment this line, it is used in the perform_data_check function
# workbook = None

def open_file_dialog():
    global gv_file_path, sheet_frame, message_frame, status_text
    # If sheet_frame is not None, destroy it to clean up the previous sheet buttons
    if sheet_frame is not None:
        sheet_frame.destroy()
        sheet_frame = None
    # If message_frame is not None, destroy it to clean up the previous messages
    if message_frame is not None:
        message_frame.destroy()
        message_frame = None
    # # clear the status 
    # if status_text is not None:
    #     status_text.delete(1.0, tk.END)
    gv_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    file_path_label.config(text=gv_file_path)
    populate_sheet_buttons()  # Call the function to populate sheet buttons after file selection

sheet_frame = None
message_frame = None

# todo - make the workbook and worksheet global variables, no need to open the file each time
def populate_sheet_buttons():
    global sheet_frame, message_frame, message_text, status_text, workbook
    
    workbook = openpyxl.load_workbook(gv_file_path,data_only=True,keep_links=False)
    sheet_names = workbook.sheetnames
    
    # Create sheet frame to hold sheet buttons
    if sheet_frame is None:
        sheet_frame = tk.Frame(root, bg="white", bd=0, relief="solid")
        sheet_frame.pack(side=tk.TOP, anchor=tk.W)
    
    for sheet_name in sheet_names:
        if not workbook[sheet_name].sheet_state == 'hidden':
            sheet_button = tk.Button(sheet_frame, text=sheet_name, command=lambda sheet=sheet_name: perform_data_check(sheet))
            sheet_button.pack(side=tk.LEFT)  # Align buttons horizontally
    
    # Create message frame to display the messages and results 
    if message_frame is None:
        message_frame = tk.Frame(root, bg="white", bd=1, relief="solid")
        message_frame.pack(side=tk.TOP, anchor=tk.W, fill=tk.BOTH, expand=True)
    
    # Create text area to display messages and results
    message_text = tk.Text(message_frame, height=100, width=200, font=("Arial", 10), fg="grey")
    message_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # Adjust the fill parameter to fill both horizontally and vertically

    # Create status text area to display the number of errors
    if status_text is None:
        status_text = tk.Text(status_frame, height=1, width=500, font=("Arial", 10), fg="blue")
        status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)


    # Create vertical scroll bar
    scroll_bar = tk.Scrollbar(message_frame)
    scroll_bar.pack(side=tk.RIGHT, fill=tk.Y)

    # Configure text area to use vertical scroll bar
    message_text.config(yscrollcommand=scroll_bar.set)
    scroll_bar.config(command=message_text.yview)

    # Initialize the text area with a welcome message
    message_text.insert(tk.END, "Welcome to SUREINFO SAP Data Migration Check Tool!\n")
    message_text.insert(tk.END, "Please select an Excel file to perform data check.\n\n\n")
    message_text.insert(tk.END, "*** DISCLAMER ***\n\n")
    message_text.insert(tk.END, f"The SUREINFO Data Migration Check Tool is provided under license by Corporate Name. The Tool is intended solely for use by authorized personnel and is governed by the terms and conditions of the licensing agreement between the Company and the licensed user.\n\n")
    message_text.insert(tk.END, f"The Tool is designed to assist in the migration of data, and while every effort has been made to ensure its accuracy and reliability, the Company makes no warranties, express or implied, regarding the Tools functionality or its fitness for any particular purpose. The Licensee is responsible for ensuring that the data migration complies with all applicable laws and regulations.\n\n")
    message_text.insert(tk.END, f"The Company shall not be liable for any direct, indirect, incidental, consequential, or punitive damages arising out of the use or inability to use the Tool. The Licensee agrees to indemnify and hold harmless the Company from any claims, damages, losses, or expenses incurred as a result of data migration activities.\n\n")
    message_text.insert(tk.END, f"Use of the Tool constitutes acceptance of these terms. If you do not agree to these terms, do not use the Tool.\n")
    message_text.insert(tk.END, f"Licnese Expiry Date: {gv_licesne_expire_date.strftime('%d %b %Y')}\n")
    # Insert the file size and number of sheets into the status text
    status_text.delete(1.0, tk.END)
    status_text.insert(tk.END, f"File Size: {round((open(gv_file_path, 'rb').seek(0, 2) / 1024), 2)} KB, Sheets: {len(sheet_names)}\n")    


def is_number(s):
    if s is None:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False

def perform_data_check(sheet_name):
    # Perform data check logic for the selected sheet
    # messagebox.showinfo("Data Check", f"Performing data check for sheet: {sheet_name}")
    # mandatory check: check for each column, data since row 8 to the last row, if row 6 is 'R' and the cell is empty, then show error message in the text area


    #clear the text area
    message_text.delete(1.0, tk.END)
    # clear the status text
    status_text.delete(1.0, tk.END) # type: ignore

    # workbook = openpyxl.load_workbook(gv_file_path) #it works only if not read only

    # get the sheet
    sheet = workbook[sheet_name]
    # check if the file is a correct template, A3 should be 'TABLE' and A4 should be 'FIELD'
    if sheet.cell(row=3, column=1).value != 'TABLE' or sheet.cell(row=4, column=1).value != 'FIELD':
        status_text.insert(tk.END, "Invalid template, please select the correct template.\n") # type: ignore
        return

    # get the last row
    last_row = sheet.max_row
    while sheet.cell(row=last_row, column=2).value is None:
        last_row -= 1
    # get the last column
    last_column = sheet.max_column
    while sheet.cell(row=2, column=last_column).value is None:
        last_column -= 1
    # get the reference row
    reference_row = tuple(sheet[1])
    # get the header row
    header_row = tuple(sheet[2])
    # get the data type row
    data_type_row = tuple(sheet[5])
    # get the mandatory row
    mandatory_row = tuple(sheet[6])
    # get the data rows
    data_rows = tuple(sheet[1:last_row])

    # initialize a reference check list, and loop the reference row to get the reference sheet. 
    reference_check_list = []
    # loop the reference row to get the reference sheet, and read the reference sheet to get the reference data with the same field name
    for i in range(1, last_column):
        reference_sheet_name = reference_row[i].value
        reference_field_name = header_row[i].value
        if reference_sheet_name is not None and reference_field_name is not None and reference_sheet_name in workbook.sheetnames:
            reference_sheet = workbook[str(reference_sheet_name)]
            # get the column index of the reference field in the reference sheet
            reference_field_column = None
            for cell in reference_sheet[2]:
                if cell.value == reference_field_name:
                    reference_field_column = cell.column
                    break

            # if the reference field is found in the reference sheet, get the reference data. change min_row to 3 to adapt different format of ref data
            if reference_field_column is not None:
                reference_data = list(reference_sheet.iter_cols(min_row=3, max_row=reference_sheet.max_row, min_col=reference_field_column, max_col=reference_field_column))
            else:
                reference_data = []

            reference_check_list.append((reference_sheet_name, reference_field_name, reference_data))

    # show the progress bar
    progress_bar['value'] = 0
    progress_bar['maximum'] = last_row

    # loop each columns and rows to check the mandatory fields and data length
    for i in range(7, last_row):
        # update the progress bar
        progress_bar['value'] += 1
        root.update_idletasks()

        # create a list to hold the required fields
        check_result = []
        for j in range(1, last_column):
            field_name = header_row[j].value
            cell_value = data_rows[i][j].value # type: ignore
            data_type = data_type_row[j].value
            mandatory_indicator = mandatory_row[j].value

            # if mandatory_inictor is null or empty, set it to 'N'
            if mandatory_indicator is None or str(mandatory_indicator).strip() == '':
                mandatory_indicator = 'N'

            # check the mandatory fields
            if mandatory_indicator.upper()[0] in ['R', 'F'] and (cell_value is None or str(cell_value).strip() == ''):  # type: ignore
                check_result.append(f" \"{field_name}\" is required;")

            # check the data length if data type is CHAR
            if data_type is not None and data_type.startswith('CHAR') and cell_value is not None: # type: ignore
                try:
                    data_length = int(data_type.split('(')[1].split(')')[0].split(',')[0]) # type: ignore
                except Exception as e:
                    print(f"245 mandatory check {e}")
                    data_length = 500       
                # cell_value = int(cell_value) if is_number(cell_value) else cell_value # dont change orginal field value
                #if len(str(int(cell_value) if is_number(cell_value) else cell_value)) > data_length:

                try:
                    if isinstance(cell_value, (int, float)):
                        cell_value_lencheck = int(cell_value) if isinstance(cell_value, float) else cell_value
                        if len(str(cell_value_lencheck)) > data_length:
                            check_result.append(f" \"{field_name}\" length {len(str(cell_value_lencheck))} exceeds {data_length};")
                    elif len(str(cell_value)) > data_length:
                        check_result.append(f" \"{field_name}\" length {len(str(cell_value))} exceeds {data_length};")
                except Exception as e:
                    print(f"264 data type check:  {e}")


                # if len(str((cell_value) if is_number(cell_value) else cell_value)) > data_length:
                #     check_result.append(f" \"{field_name}\" length {len(str(cell_value))} exceeds {data_length};")


            # if the data type is number, check the number format
            elif data_type is not None and (data_type.startswith('NUM') or data_type.startswith('CURR') or data_type.startswith('QUA') or data_type.startswith('DEC') ) and cell_value is not None and not (str(cell_value).isnumeric() or str(cell_value).replace('.', '', 1).isdigit()): # type: ignore
                check_result.append(f" \"{field_name} = {str(cell_value)}\" is not a number or digit;")

            # if the data type is date, check the if date format is correct
            elif data_type is not None and data_type.startswith('DAT') and cell_value is not None: # type: ignore
    
                # check date string in format YYYYMMDD
                if len(str(cell_value)) == 8:
                    try:
                        datetime.datetime.strptime(str(cell_value), '%Y%m%d')
                    except Exception as e:
                        print(e)
                        check_result.append(f" \"{field_name}\" date format incorret;")

                else:
                    check_result.append(f" \"{field_name}\" date format incorret;")

            reference_row_value = reference_row[j].value
            # if the reference field is not empty, check the reference data
            if reference_row_value is not None and cell_value is not None: # type: ignore
                # filter the reference_check_list based on reference_row[j] and field_name
                try:
                    filtered_check_list = [item[2][0] for item in reference_check_list if item[0] == reference_row_value and item[1] == field_name]
                except Exception as e:
                    filtered_check_list = []

                if filtered_check_list:
                    # check if cell_value is in the reference data of the filtered list
                    # if cell_value not in filtered_check_list[0]: # type: ignore
                    if str(cell_value) not in [str(item.value) for item in filtered_check_list[0]]:
                        check_result.append(f" \"{field_name} = {str(cell_value)}\" is not found in {reference_row_value};")
                
                # duplicate check, count of same values since row 7 to i-1 should less than 1
                if str(reference_row_value).upper() in ['DUP', 'KEY']: 
                    if [data_rows[x][j].value for x in range(7, i)].count(cell_value) > 0: # type: ignore
                        check_result.append(f" \"{field_name} = {str(cell_value)}\" is duplicated;")
                #[data_rows[x][j].value for x in range(7, last_row)]: # type: ignore
            
        
                
        # insert the error message or success message into the text area
        if check_result:
            date = datetime.datetime.now().strftime('%d %b')
            error_message = f"{i+1} \t{date} SAP: {' '.join(check_result)}"
            message_text.insert(tk.END, error_message + "\n")
        else:
            record = f"{i+1} \t - "
            message_text.insert(tk.END, record + "\n")
            # show no technical data errors in the status frame
    
    # hide the progress bar
    progress_bar['value'] = 0
    

    # show the number of errors in the status frame
    status_text.delete(1.0, tk.END) # type: ignore
    error_count = message_text.get(1.0, tk.END).count('SAP:')
    if error_count > 0:
        status_text.insert(tk.END, f"Template No: {sheet_name}, Number of Errors: {error_count}") # type: ignore
    else:
        status_text.insert(tk.END, "No data errors") # type: ignore
    
    # close the workbook todo
    # workbook.close()

def copy_to_clipboard():
    if message_text is None:
        return
    
    root.clipboard_clear()
    root.clipboard_append(message_text.get(1.0, tk.END))
    root.update()  # now it stays on the clipboard after the window is closed
    # update status text with orginal message by adding (copied) in the tail if the copy was successful

    status_text.delete(1.0, tk.END) # type: ignore
    status_text.insert(tk.END,"Result data copied.") # type: ignore


# Create the main window
root = tk.Tk()
root.geometry("1024x768")
root.configure(bg="light grey")
root.title(f"{gv_title} {gv_version}")
# change my icon - needs resource file during packing
# root.iconbitmap('favicon.ico')


#check license if the date is bigger than 2024-01-01, then show the lincese expired message, confirm and close the application
if datetime.datetime.now() > gv_licesne_expire_date:
    messagebox.showinfo("License Expired", "The license for this application has expired. Please contact the vendor for renewal.")
    root.destroy()


# Create title area with orange background
title_frame = tk.Frame(root, bg="orange")
title_frame.pack(fill=tk.X)

# Create title label
title_label = tk.Label(title_frame, text=f"{gv_title_label} {gv_version}", font=("Arial", 16), bg="orange", fg="white")
title_label.pack(pady=10)
# Create a frame to hold the button and label
button_frame = tk.Frame(root, bg="white")
button_frame.pack(side=tk.TOP, anchor=tk.NW, fill=tk.X, expand=True)

# Create button to open file dialog
open_button = tk.Button(button_frame, text=gv_button_label, command=open_file_dialog)
open_button.pack(side=tk.LEFT, padx=10, pady=10)

# Create label to display file path
file_path_label = tk.Label(button_frame, text="...select the file...", font=("Arial", 10), bg="white")
file_path_label.pack(side=tk.LEFT, padx=10, pady=10)

# create button to copy the result to clipboard
copy_button = tk.Button(button_frame, text="Copy to Clipboard", command=copy_to_clipboard)
copy_button.pack(side=tk.RIGHT, padx=10, pady=10)


gv_file_path = ""

# create a frame to hold the status message in the bottom
status_frame = tk.Frame(root, bg="white")
status_frame.pack(side=tk.BOTTOM, anchor=tk.W)


# Create a frame to hold the progress bar
progress_frame = tk.Frame(root, bg="white")
progress_frame.pack(side=tk.BOTTOM, anchor=tk.W)
progress_bar = ttk.Progressbar(progress_frame, length=2000, mode='determinate')
progress_bar.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

root.mainloop()