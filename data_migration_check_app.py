# Streamlit Data Migration Check Tool
# This is a starter template for your new Streamlit app.
# You can migrate your logic from the reference code in ref_code/SUREINFO_SAP_DATA_CHECK.py here.

import streamlit as st
import pandas as pd
import openpyxl
import datetime
import base64
import io

st.set_page_config(layout="wide")

APP_TITLE = "SUREINFO SAP S/4 Data Migration Cleansing Tool"
APP_VERSION = "Version 1.0.7"
LICENSE_EXPIRE_DATE = datetime.datetime(2025, 12, 31)
EXCEL_GREEN = "#217346"
NAV_BG = "#f7f7f7"

st.markdown(f"<h1 style='margin-bottom:0'>{APP_TITLE} <span style='font-size:0.7em;color:gray'>{APP_VERSION}</span></h1>", unsafe_allow_html=True)

if datetime.datetime.now() > LICENSE_EXPIRE_DATE:
    st.error("The license for this application has expired. Please contact the vendor for renewal.")
    st.stop()

# Collapsible navigation area
if 'nav_expanded' not in st.session_state:
    st.session_state['nav_expanded'] = True

nav_col, work_col = st.columns([
    0.25 if st.session_state['nav_expanded'] else 0.05,
    0.75 if st.session_state['nav_expanded'] else 0.95
], gap="small")

with nav_col:
    st.markdown(f"<div style='background:{NAV_BG};padding:0.5em 0.5em 0.5em 0.5em;border-radius:8px;display:flex;flex-direction:column;'>", unsafe_allow_html=True)
    # Navigation content (always show collapse/expand at bottom)
    st.markdown(f"<div style='background:{EXCEL_GREEN};color:white;padding:0.5em 1em;border-radius:6px;margin-bottom:1em;font-weight:bold;'>Navigation</div>", unsafe_allow_html=True)
    if st.session_state['nav_expanded']:
        st.info("Upload an Excel file and select a sheet to check.")
        excel_file = st.file_uploader("Select Excel File", type=["xlsx"], key="excel_upload")
        if excel_file:
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True, keep_links=False)
                visible_sheets = [s for s in wb.sheetnames if wb[s].sheet_state != 'hidden']
                st.session_state['sheetnames'] = visible_sheets
                st.session_state['wb_bytes'] = excel_file.getvalue()
            except Exception as e:
                st.error(f"Failed to load Excel file: {e}")
                st.stop()
            st.success(f"Loaded file with {len(visible_sheets)} visible sheets.")
            for sheet in visible_sheets:
                btn = st.button(sheet, key=f"sheet_btn_{sheet}", help=f"Check {sheet}",
                    use_container_width=True,
                    type="secondary")
                if btn:
                    st.session_state['selected_sheet'] = sheet
                    st.session_state['run_check'] = True
                    st.rerun()
        else:
            st.session_state['sheetnames'] = []
            st.session_state['selected_sheet'] = None
            st.session_state['run_check'] = False
        st.markdown("<div style='flex:1'></div>", unsafe_allow_html=True)
    else:
        st.markdown("<div style='flex:1'></div>", unsafe_allow_html=True)
    # Collapse/Expand button always at bottom
    st.divider()
    nav_toggle = st.button(
        "⏪" if st.session_state['nav_expanded'] else "⏩",
        key="nav_toggle_btn",
        help="Expand/collapse navigation area"
    )
    if nav_toggle:
        st.session_state['nav_expanded'] = not st.session_state['nav_expanded']
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

with work_col:
    st.markdown(f"<div style='background:#f4f4f4;padding:0.5em 1em;border-radius:6px;margin-bottom:1em;font-weight:bold;'>Working Area</div>", unsafe_allow_html=True)
    if st.session_state.get('selected_sheet') and st.session_state.get('wb_bytes') and st.session_state.get('run_check'):
        wb = openpyxl.load_workbook(io.BytesIO(st.session_state['wb_bytes']), data_only=True, keep_links=False)
        sheet = st.session_state['selected_sheet']
        ws = wb[sheet]
        st.markdown(f"### Checking Sheet: <span style='color:{EXCEL_GREEN}'>{sheet}</span>", unsafe_allow_html=True)
        if ws.cell(row=3, column=1).value != 'TABLE' or ws.cell(row=4, column=1).value != 'FIELD':
            st.error("Invalid template, please select the correct template.")
        else:
            last_row = ws.max_row
            while ws.cell(row=last_row, column=2).value is None and last_row > 1:
                last_row -= 1
            last_col = ws.max_column
            while ws.cell(row=2, column=last_col).value is None and last_col > 1:
                last_col -= 1
            reference_row = tuple(ws[1])
            header_row = tuple(ws[2])
            data_type_row = tuple(ws[5])
            mandatory_row = tuple(ws[6])
            data_rows = tuple(ws[1:last_row])
            reference_check_list = []
            for i in range(1, last_col):
                reference_sheet_name = reference_row[i].value
                reference_field_name = header_row[i].value
                if reference_sheet_name and reference_field_name and reference_sheet_name in wb.sheetnames:
                    ref_ws = wb[reference_sheet_name]
                    ref_col = None
                    for cell in ref_ws[2]:
                        if cell.value == reference_field_name:
                            ref_col = cell.column
                            break
                    if ref_col:
                        ref_data = list(ref_ws.iter_cols(min_row=3, max_row=ref_ws.max_row, min_col=ref_col, max_col=ref_col))
                    else:
                        ref_data = []
                    reference_check_list.append((reference_sheet_name, reference_field_name, ref_data))
            results = []
            error_count = 0
            progress = st.progress(0)
            for i in range(7, last_row):
                check_result = []
                for j in range(1, last_col):
                    field_name = header_row[j].value
                    cell_value = data_rows[i][j].value
                    data_type = data_type_row[j].value
                    mandatory_indicator = mandatory_row[j].value
                    if mandatory_indicator is None or str(mandatory_indicator).strip() == '':
                        mandatory_indicator = 'N'
                    if mandatory_indicator.upper()[0] in ['R', 'F'] and (cell_value is None or str(cell_value).strip() == ''):
                        check_result.append(f' "{field_name}" is required;')
                    if data_type and str(data_type).startswith('CHAR') and cell_value is not None:
                        try:
                            data_length = int(str(data_type).split('(')[1].split(')')[0].split(',')[0])
                        except Exception:
                            data_length = 500
                        try:
                            if isinstance(cell_value, (int, float)):
                                cell_value_lencheck = int(cell_value) if isinstance(cell_value, float) else cell_value
                                if len(str(cell_value_lencheck)) > data_length:
                                    check_result.append(f' "{field_name}" length {len(str(cell_value_lencheck))} exceeds {data_length};')
                            elif len(str(cell_value)) > data_length:
                                check_result.append(f' "{field_name}" length {len(str(cell_value))} exceeds {data_length};')
                        except Exception:
                            pass
                    elif data_type and (str(data_type).startswith('NUM') or str(data_type).startswith('CURR') or str(data_type).startswith('QUA') or str(data_type).startswith('DEC')) and cell_value is not None and not (str(cell_value).isnumeric() or str(cell_value).replace('.', '', 1).isdigit()):
                        check_result.append(f' "{field_name} = {str(cell_value)}" is not a number or digit;')
                    elif data_type and str(data_type).startswith('DAT') and cell_value is not None:
                        if len(str(cell_value)) == 8:
                            try:
                                datetime.datetime.strptime(str(cell_value), '%Y%m%d')
                            except Exception:
                                check_result.append(f' "{field_name}" date format incorret;')
                        else:
                            check_result.append(f' "{field_name}" date format incorret;')
                    reference_row_value = reference_row[j].value
                    if reference_row_value and cell_value is not None:
                        try:
                            filtered_check_list = [item[2][0] for item in reference_check_list if item[0] == reference_row_value and item[1] == field_name]
                        except Exception:
                            filtered_check_list = []
                        if filtered_check_list:
                            if str(cell_value) not in [str(item.value) for item in filtered_check_list[0]]:
                                check_result.append(f' "{field_name} = {str(cell_value)}" is not found in {reference_row_value};')
                        if str(reference_row_value).upper() in ['DUP', 'KEY']:
                            if [data_rows[x][j].value for x in range(7, i)].count(cell_value) > 0:
                                check_result.append(f' "{field_name} = {str(cell_value)}" is duplicated;')
                if check_result:
                    error_count += 1
                    date = datetime.datetime.now().strftime('%d %b')
                    error_message = f"{i+1} \t{date} SAP: {' '.join(check_result)}"
                    results.append(error_message)
                else:
                    record = f"{i+1} \t - "
                    results.append(record)
                progress.progress((i-6)/(last_row-7))
            progress.empty()
            st.subheader("Check Results")
            result_text = '\n'.join(results)
            st.text_area("Result", result_text, height=400)
            st.info(f"Sheet: {sheet}, Number of Errors: {error_count}")
            b64 = base64.b64encode(result_text.encode()).decode()
            href = f'<a href="data:file/txt;base64,{b64}" download="check_result.txt">Download Result as TXT</a>'
            st.markdown(href, unsafe_allow_html=True)
    else:
        st.info("Please select a sheet from the navigation area.")
if not st.session_state.get('wb_bytes'):
    st.info("Please upload an Excel file in the navigation area.")
